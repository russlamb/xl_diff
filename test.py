"""
This module contains unit tests for the comparison library
"""
import re
import unittest
from datetime import datetime

import openpyxl as xl
import csv
from itertools import zip_longest
from xl_diff import compare_files, SqlCompare, SummaryNode, summarize_differences, write_summary_file, is_number, \
    is_date, ValueNode, make_sorted_sheet, sort_values, SqlToXl, process_file,convert_csv_to_excel
from dateutil.parser import parse
import os

from sql_compare import sql_compare_configure_arg_parser
from compare import compare_excel_configure_arg_parser
from sql_compare_file import sql_compare_file_configure_arg_parser

# logging.basicConfig(level=logging.DEBUG, format='%(asctime)-15s %(message)s')

TEST_DB_OUTPUT_XLSX = r".\test_db\output.xlsx"

TEST_DB_DEMO_DB = r".\test_db\demo.db"
TEST_DB_CONNECTION_STRING = "Driver={SQLite3 ODBC Driver};Database=%s;Version=3" % TEST_DB_DEMO_DB

TEST_DB_RIGHT_XLSX = r".\test_db\right.xlsx"

TEST_DB_LEFT_XLSX = r".\test_db\left.xlsx"

TESTS_SUMMARY_XLSX = r"tests\summary.xlsx"

TESTS_OUTPUT2_XLSX = r"tests\output2.xlsx"

TESTS_RIGHT2_CSV = r"tests\right2.csv"

TESTS_LEFT_CSV = r"tests\left.csv"

TESTS_OUTPUT_XLSX = r"tests\output.xlsx"

TESTS_RIGHT_XLSX = r"tests\right.xlsx"

TESTS_LEFT_XLSX = r"tests\left.xlsx"

TEST_DB_RIGHT2_XLSX = r"test_db\right2.xlsx"

TESTS_OUTPUT_MULTI_XLSX = r"test_db\output_multithreaded.xlsx"

TEST_DB_RIGHT_MULTI_XLSX = r".\test_db\right_multithreaded.xlsx"

TEST_DB_LEFT_MULTI_XLSX = r".\test_db\left_multithreaded.xlsx"

TESTS_CMD_OUTPUT_XLSX = r"test_cmd_line\output.xlsx"
TESTS_CMD_LEFT_XLSX = r"test_cmd_line\left.xlsx"
TESTS_CMD_RIGHT_XLSX = r"test_cmd_line\right.xlsx"


class TestConvert(unittest.TestCase):
    """
    Test converting CSV file to Excel
    """

    def test_convert_csv(self):
        csv_path = TESTS_LEFT_CSV
        excel_file = convert_csv_to_excel(csv_path)
        wb = xl.load_workbook(excel_file)
        ws = wb.active
        with open(csv_path, newline='') as csv_file:
            reader = csv.reader(csv_file, delimiter=',', quotechar='"')
            cells_checked = 0
            for (excel_row, csv_row) in zip_longest(ws.rows, reader):
                for i in range(0, len(excel_row)):
                    self.assertEqual(excel_row[i].value, csv_row[i], "excel value differs from csv")
                    cells_checked += 1
            self.assertGreater(cells_checked, 0)  # greater than zero


class TestExcel(unittest.TestCase):
    """
    Test sorting, and comparison of files
    """

    def setUp(self):

        self.left_xlsx = TESTS_LEFT_XLSX
        self.right_xlsx = TESTS_RIGHT_XLSX
        self.left_wb = xl.load_workbook(self.left_xlsx)
        self.right_wb = xl.load_workbook(self.right_xlsx)

        self.left_sheet = self.left_wb.worksheets[0]
        self.right_sheet = self.right_wb.worksheets[0]
        self.val = sort_values(self.left_sheet, self.right_sheet, 1, True)

    def test_sort(self):
        """
        Using a known dataset, run comparison and compare the output with the expected result
        :return: None
        """
        expected_values = [(2, 2, "Row 1"), (3, 3, "Row 2"), (5, None, "Row 3"), (4, 4, "Row 4")]
        test_list = [ValueNode(f, r, v) for (f, r, v) in expected_values]
        check_count = 0
        for (i, x) in zip_longest(self.val, test_list):
            self.assertEqual(i, x, "test value does not match expected value")
            check_count += 1
        self.assertGreater(check_count, 0)  # greater than zero

    def test_make_sorted_sheet(self):
        """
        Using a known input, run sort and compare to expected result
        :return: None
        """
        wb = xl.Workbook()
        sheet_name = self.left_wb.sheetnames[0]
        left_sheet_name = 'left_' + sheet_name
        left_sheet = make_sorted_sheet(wb, self.left_sheet, self.val, left_sheet_name, 'left', True)
        expected_values = [(1, 1, "Header"), (2, 2, "Row 1"), (3, 3, "Row 2"), (5, None, "Row 3"), (4, 4, "Row 4")]
        check_count = 0
        for row in range(1, left_sheet.max_row + 1):
            self.assertEqual(left_sheet.cell(row=row, column=1).value, expected_values[row - 1][2])
            check_count += 1
        self.assertGreater(check_count, 0)  # greater than zero

    def test_compare_files_xlsx(self):
        """
        Using a known input, run comparison and compare to expected result
        :return: None
        """
        compare_files(self.left_xlsx, self.right_xlsx, TESTS_OUTPUT_XLSX, open_on_finish=False,
                      sort_column=1, compare_type="sorted", sheet_matching="order")
        expected_values = [
            ["Header", "Header", "Same", "Col A", "Col A", "Same", "Col B", "Col B", "Same"],
            ["Row 1", "Row 1", "Same", "1", 1, 0, "2", 3, 1],
            ["Row 2", "Row 2", "Same", "z", "z", "Same", "Q", "W", "Different"],
            ["Row 3", None, "Different", "extra", None, "Different", "row", None, "Different"],
            ["Row 4", "Row 4", "Same", "1/1/2019", datetime(2019, 1, 1), "Different", "2/2/2012",
             datetime(2012, 2, 1), "Different"]
        ]

        wb = xl.load_workbook(TESTS_OUTPUT_XLSX)
        ws = wb.worksheets[2]  # third worksheet is diff
        for col in range(1, ws.max_column + 1):
            for row in range(1, ws.max_row + 1):

                value = ws.cell(row=row, column=col).value
                expected_value = expected_values[row - 1][col - 1]
                if is_number(value) and is_number(expected_value):
                    self.assertEqual(float(value), float(expected_value))
                else:
                    self.assertEqual(str(value), str(expected_value))

    def test_compare_files_csv(self):
        """
        Using a known input, convert csv to excel and compare to expected result
        :return: None
        """
        compare_files(TESTS_LEFT_CSV, TESTS_RIGHT2_CSV, TESTS_OUTPUT2_XLSX, open_on_finish=False,
                      sort_column=1, compare_type="sorted", sheet_matching="order")
        expected_values = [
            ["Header", "Header", "Same", "Col A", "Col A", "Same", "Col B", "Col B", "Same"],
            ["Row 1", "Row 1", "Same", "1", 1, 0, "2", 3, 1],
            ["Row 2", "Row 2", "Same", "z", "z", "Same", "Q", "W", "Different"],
            ["Row 3", None, "Different", "extra", None, "Different", "row", None, "Different"],
            ["Row 4", "Row 4", "Same", "1/1/2019", "1/1/2019", "Same", "2/2/2012",
             "2/1/2012", "Different"]
        ]
        wb = xl.load_workbook(TESTS_OUTPUT2_XLSX)
        ws = wb.worksheets[2]  # third worksheet is diff
        for col in range(1, ws.max_column + 1):
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                expected_value = expected_values[row - 1][col - 1]
                # logging.debug("row {}, col {}: {} vs {}".format(row, col, value, expected_value))
                if is_number(value) and is_number(expected_value):
                    self.assertEqual(float(value), float(expected_value))
                elif is_date(value) and is_date(expected_value):
                    self.assertEqual(parse(value), parse(expected_value))
                else:
                    self.assertEqual(str(value), str(expected_value))

    def test_sort_column_list_one_value(self):
        """
        
        :return:
        """
        expected_values = [(2, 2, "Row 1"), (3, 3, "Row 2"), (5, None, "Row 3"), (4, 4, "Row 4")]
        test_list = [ValueNode(f, r, (v,)) for (f, r, v) in expected_values]
        check_count = 0
        sorted_values = sort_values(self.left_sheet, self.right_sheet, [1], True)

        for (i, x) in zip_longest(sorted_values, test_list):
            self.assertEqual(i, x, "test value does not match expected value")
            check_count += 1
        self.assertGreater(check_count, 0)  # greater than zero

    def test_sort_multiple_columns(self):
        expected_values = [(2, 2, ('Row 1', '1')), (3, 3, ('Row 2', 'z')), (5, None, ('Row 3', 'extra')),
                           (4, None, ('Row 4', '1/1/2019')), (None, 4, ('Row 4', '2019-01-01 00:00:00'))]
        test_list = [ValueNode(f, r, v) for (f, r, v) in expected_values]
        check_count = 0
        sorted_values = sort_values(self.left_sheet, self.right_sheet, [1, 2], True)

        for (i, x) in zip_longest(sorted_values, test_list):
            self.assertEqual(i, x, "test value does not match expected value")
            check_count += 1
        self.assertGreater(check_count, 0)  # greater than zero


class TestSummary(unittest.TestCase):
    """
    Test the summary module
    """

    def setUp(self):
        self.output_xlsx = TESTS_OUTPUT_XLSX
        self.output_wb = xl.load_workbook(self.output_xlsx)
        self.output_sheet = self.output_wb.worksheets[2]
        self.summary_xlsx = TESTS_SUMMARY_XLSX

    def test_get_values(self):
        nodes = summarize_differences(self.output_sheet, 1, 3, 0.001, True)
        self.assertEqual(len(nodes), 3, "Wrong number of columns")

        def check_node(n: SummaryNode, column_name, expected_differences):
            self.assertEqual(n.column_with_differences, column_name, "Wrong column name")
            self.assertEqual(n.number_of_differences, expected_differences, "Wrong number of differences")

        check_node(nodes[0], "Header", 1)  # Header has 1 diference
        check_node(nodes[1], "Col A", 2)  # Col A has 2 differences
        check_node(nodes[2], "Col B", 4)  # Col B has 4 differecnes

    def test_summary_file(self):
        write_summary_file(self.output_xlsx, self.summary_xlsx)
        summary_wb = xl.load_workbook(self.summary_xlsx)
        summary_sheet = summary_wb["summary"]

        data = [
            ["Sheet Name", "Column Name", "Number of Differences", "Total Rows", "Percent Different", "Column Index"]
            , ["Sheet v Sheet1", "Header", 1, 5, "20.00%", 1]
            , ["Sheet v Sheet1", "Col A", 2, 5, "40.00%", 2]
            , ["Sheet v Sheet1", "Col B", 4, 5, "80.00%", 3]
        ]
        for i in range(1, summary_sheet.max_row + 1):
            for j in range(1, summary_sheet.max_column + 1):
                cell_value = summary_sheet.cell(row=i, column=j).value
                expected_value = data[i - 1][j - 1]
                self.assertEqual(cell_value, expected_value,
                                 "cell had {} expected {}".format(cell_value, expected_value))


class TestSqlToXl(unittest.TestCase):
    """
    Test the SQl to Xl class

    Note: there may be some issues with running the test_save_to_file function on its own, but running the whole
    class seems to succeed.  Possibly a connection issue.
    """

    def setUp(self):
        self.s2x = SqlToXl(TEST_DB_CONNECTION_STRING)

    def test_query_table(self):
        rows = self.s2x.get_query_results("select * from left")
        self.assertEqual(len(rows), 4, "Wrong number of rows in test database")
        print(rows)

    def test_save_to_file(self):
        target_path = TEST_DB_LEFT_XLSX
        self.s2x.save_sql("select * from left", target_path)
        self.assertTrue(os.path.exists(target_path), "File was not created")
        os.remove(target_path)
        self.assertFalse(os.path.exists(target_path), "File was not deleted")


class TestSqlCompare(unittest.TestCase):
    """
    Test the SQL to Compare class

    """

    def setUp(self):
        self.same_db = SqlCompare(TEST_DB_CONNECTION_STRING,
                                  TEST_DB_CONNECTION_STRING,
                                  TEST_DB_LEFT_XLSX,
                                  TEST_DB_RIGHT_XLSX)
        self.multithreaded = SqlCompare(TEST_DB_CONNECTION_STRING,
                                        TEST_DB_CONNECTION_STRING,
                                        TEST_DB_LEFT_MULTI_XLSX,
                                        TEST_DB_RIGHT_MULTI_XLSX,
                                        multithreaded=True)

    def test_generate_files_single_query(self):
        """
        Check the SqlCompare class can generate two files from a single query.
        :return: None
        """
        (left_path, right_path) = self.same_db.generate_files_from_query("select * from left")
        self.assertTrue(os.path.exists(left_path), "Left file was not generated")
        self.assertTrue(os.path.exists(right_path), "Right file was not generated")
        os.remove(left_path)
        self.assertFalse(os.path.exists(left_path), "Left file was not removed")
        os.remove(right_path)
        self.assertFalse(os.path.exists(right_path), "Right file was not removed")

    def test_generate_files_two_queries(self):
        """
        Check the sql compare class can generate two files from two different queries
        :return: None
        """
        (left_path, right_path) = self.same_db.generate_files_from_query("select * from left", "select * from right2")
        self.assertTrue(os.path.exists(left_path), "Left file was not generated")
        self.assertTrue(os.path.exists(right_path), "Right file was not generated")
        os.remove(left_path)
        self.assertFalse(os.path.exists(left_path), "Left file was not removed")
        os.remove(right_path)
        self.assertFalse(os.path.exists(right_path), "Right file was not removed")

    def test_compare_two_identical_files(self):
        """
        Generate two files from the same database using the same query, then compare them
        :return: None
        """
        output_path = self.same_db.compare_query_results(TEST_DB_OUTPUT_XLSX, "select * from left")
        self.assertTrue(os.path.exists(output_path), "Output file was not generated")
        os.remove(output_path)
        self.assertFalse(os.path.exists(output_path), "Output file was not removed")

    def test_compare_two_different_files(self):
        """
        Generate two files using two different queries, then compare them
        :return:
        """
        output_path = self.same_db.compare_query_results(TEST_DB_OUTPUT_XLSX, "select * from left",
                                                         "select * from right2")
        self.assertTrue(os.path.exists(output_path), "Output file was not generated")
        os.remove(output_path)
        self.assertFalse(os.path.exists(output_path), "Output file was not removed")

    def test_multithread_compare(self):
        """
        Use the multithreaded version of sql compare.  This will run both sql commands simultaneously.
        :return:
        """
        output_path = self.multithreaded.compare_query_results(TESTS_OUTPUT_MULTI_XLSX, "select * from left",
                                                               "select * from right2")
        self.assertTrue(os.path.exists(output_path), "Output file was not generated")
        os.remove(output_path)
        self.assertFalse(os.path.exists(output_path), "Output file was not removed")
        os.remove(TEST_DB_LEFT_MULTI_XLSX)
        os.remove(TEST_DB_RIGHT_MULTI_XLSX)

    def test_multithread_compare_error(self):
        """
        Test what happens when the threaded version of sql comparison encounters an error
        :return: None
        """
        try:
            output_path = self.multithreaded.compare_query_results(TESTS_OUTPUT_MULTI_XLSX, "select * from left",
                                                                   "select * from NOT_A_TABLE")  # select from non-existant table
        except Exception as e:
            self.assertIsInstance(e, Exception, "check the exception type")

        try:
            os.remove(TEST_DB_LEFT_MULTI_XLSX)  # remove the other file if it was created.
        except OSError:
            print("We got an OS error because the other file wasn't created.")


class TestFileProcessing(unittest.TestCase):
    """
    Check file processing functionality
    """

    def setUp(self):
        self.file = r"example_file_input\file_input.txt"

    def test_process_file(self):
        """
        Call the process file function with the test file
        :return:
        """
        process_file(True, self.file)

    def test_process_file_multithreaded(self):
        """
        Call process file function with multithreading parameter
        :return:
        """
        process_file(True, self.file, True)

    def test_process_file_compare_only(self):
        """
        Call process file function with compare only parameter
        :return:
        """
        process_file(True, self.file, True, True)

class TestArgumentParse(unittest.TestCase):
    def test_compare(self):
        parser = compare_excel_configure_arg_parser()
        cmd = f'"{TESTS_LEFT_XLSX}" "{TESTS_RIGHT_XLSX}" "{TESTS_OUTPUT_XLSX}" -c sorted -l 1 2'
        args = re.findall(r'(?:[^\s,"]|"(?:\\.|[^"])*")+', cmd)  # break apart string while respecting quotes
        x = parser.parse_args(args)
        print(x)
        self.assertEqual(x.left,f'"{TESTS_LEFT_XLSX}"')
        self.assertEqual(x.right, f'"{TESTS_RIGHT_XLSX}"')
        self.assertEqual(x.output, f'"{TESTS_OUTPUT_XLSX}"')
        self.assertEqual(x.sort_column_list, [1,2])
        self.assertEqual(x.compare_type, "sorted")


    def test_sql_compare(self):
        parser = sql_compare_configure_arg_parser()
        query = "select * from left"
        cmd = f'"{TEST_DB_CONNECTION_STRING}" "{TEST_DB_CONNECTION_STRING}" "{TEST_DB_OUTPUT_XLSX}" "{query}" -c sorted -l 1 2 -M'
        args = re.findall(r'(?:[^\s,"]|"(?:\\.|[^"])*")+', cmd)  # break apart string while respecting quotes
        x = parser.parse_args(args)
        print(x)
        self.assertEqual(x.left, f'"{TEST_DB_CONNECTION_STRING}"')
        self.assertEqual(x.right, f'"{TEST_DB_CONNECTION_STRING}"')
        self.assertEqual(x.output, f'"{TEST_DB_OUTPUT_XLSX}"')
        self.assertEqual(x.sort_column_list, [1, 2])
        self.assertEqual(x.compare_type, "sorted")

    def test_sql_compare_file(self):
        parser = sql_compare_file_configure_arg_parser()
        file = r"example_file_input\file_input.txt"
        cmd = f'"{file}"'
        args = re.findall(r'(?:[^\s,"]|"(?:\\.|[^"])*")+', cmd)
        x = parser.parse_args(args)
        print(x)
        self.assertEqual(x.file, f'"{file}"', "file paths don't match")
        self.assertFalse(x.multithreading_off, "multithreading off flag set")
        self.assertFalse(x.no_header, "no header flag set")
        self.assertFalse(x.compare_only, "compare only flag not set")

    def test_sql_compare_file_with_flags(self):
        parser = sql_compare_file_configure_arg_parser()
        file = r"example_file_input\file_input.txt"
        cmd = f'"{file}" -M -C -d'
        args = re.findall(r'(?:[^\s,"]|"(?:\\.|[^"])*")+', cmd)
        x = parser.parse_args(args)
        print(x)
        self.assertEqual(x.file, f'"{file}"', "file paths don't match")
        self.assertTrue(x.multithreading_off, "multithreading off flag set")
        self.assertTrue(x.no_header, "no header flag set")
        self.assertTrue(x.compare_only, "compare only flag not set")
