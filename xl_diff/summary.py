"""
This module is intended to summarize the output of a comparison.  It can be called independently of the comparison
module on completed excel comparisons or as part of the comparison function call itself.
"""
from collections import namedtuple

import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .validators import is_number
from .helper_excel import get_empty_workbook


class SummaryNode(namedtuple('SummaryNode', ["sheet_name", "column_with_differences", "number_of_differences",
                                             "number_of_rows", "match_percent", "column_index"])):
    """
    Contains the summarized results of a comparison.
    """


def summarize_differences(sheet, starting_column, columns_per_comparison, threshold=0.001, has_header=True,
                          diff_offset=2, starting_row=1):
    """
    Return a list of nodes containing columns with differences
    :param sheet: openpyxl worksheet object
    :param starting_column: 1 based index of first value column for "left" spreadsheet
    :param columns_per_comparison: how many columns to the right of the "left" value is next "left" value
    :param threshold: maximum numerical difference allowed
    :param has_header: if True, use the value of the first row from the "left" value column to identify the column
                        if False, use the Excel letter of the source sheet column accounting for sheets_per_comparison
    :param diff_offset: the difference column is this many columns away from the "left" value (default is 2)
    :param starting_row: one-based index of first row to check (default is 1)
    :return: list of named tuples
    """
    # these could be parameters but i made them variables
    max_col = sheet.max_column
    max_row = sheet.max_row

    summary_nodes = []  # list with output

    for col in range(starting_column, max_col + 1, columns_per_comparison):
        difference_count = 0
        number_of_rows = 0
        for row in range(starting_row, max_row + 1):
            number_of_rows += 1
            item = sheet.cell(row=row, column=col + diff_offset)
            if item.value == "Different":
                difference_count += 1
            elif is_number(item.value):
                if float(item.value) > threshold:
                    difference_count += 1
        if difference_count > 0:
            percent_different_numeric = round(difference_count / number_of_rows, 4) if number_of_rows > 0 else 0
            percent_different = "{:.2%}".format(percent_different_numeric)
            original_sheet_column = (((col - 1)  # convert one-based index to zero-based
                                      / columns_per_comparison)  # divide column index by # of cols per value
                                     + 1)  # convert back to one-based index to get the original column index
            if has_header:
                s = SummaryNode(sheet.title, sheet.cell(row=1, column=col).value, difference_count, number_of_rows,
                                percent_different, original_sheet_column)
            else:

                s = SummaryNode(sheet.title, get_column_letter(original_sheet_column), difference_count, number_of_rows,
                                percent_different, original_sheet_column)
            summary_nodes.append(s)
    return summary_nodes


def write_summary_file(input_path, output_path, sheets_per_comparison=3):
    """
    Create summary file based on comparison file.
    :param input_path: comparison file
    :param output_path: target summary file path
    :param sheets_per_comparison: number of sheets used for each sheet comparison
    :return: None
    """
    input_wb = xl.load_workbook(input_path)  # load comparison workbook
    workbook_nodes = get_workbook_nodes(sheets_per_comparison, input_wb)  # get differences for workbook
    output_wb = get_empty_workbook()
    output_wb = create_summary_worksheet(workbook_nodes, output_wb)  # create new workbook with summary info
    output_wb.save(output_path)  # save summary excel file


def get_nodes_for_workbook_path(file_path, sheets_per_comparison, starting_column=1, columns_per_comparison=3,
                                threshold=0.001, has_header=True):
    """
    Get the summary nodes for a given workbook file
    :param file_path: file path of workbook
    :param sheets_per_comparison: number of sheets in each sheet comparison.  e.g. left, right, diff
    :param starting_column: 1 based index of column to start comparison
    :param columns_per_comparison: number of sheets in each column comparison.  e.g. left, right, diff
    :param threshold:  maximum numerical difference allowed
    :param has_header: if True, use the value of the first row from the "left" value column to identify the column
                        if False, use the Excel letter of the source sheet column accounting for sheets_per_comparison
    :return: list of SummaryNodes
    """
    input_wb = xl.load_workbook(file_path)  # load comparison workbook
    return get_workbook_nodes(sheets_per_comparison, input_wb, starting_column, columns_per_comparison, threshold,
                              has_header)


def get_workbook_nodes(sheets_per_comparison, input_wb, starting_column=1, columns_per_comparison=3, threshold=0.001,
                       has_header=True):
    """
    Search comparison worksheets for differences
    :param sheets_per_comparison: number of columns used for each value comparison
    :param input_wb: openpyxl workbook object
    :return: list of tuples with summary information
    """
    workbook_nodes = []  # list of SummaryNode objects

    for i in range(sheets_per_comparison - 1, len(input_wb.worksheets), sheets_per_comparison):  # i is sheet index
        sheet = input_wb.worksheets[i]  # get comparison worksheet (at index i)
        sheet_nodes = summarize_differences(sheet, starting_column, columns_per_comparison, threshold,
                                            has_header)  # get list of nodes for the sheet
        workbook_nodes.extend(sheet_nodes)  # append sheet nodes to the end of list for the workbook

    return workbook_nodes


def create_summary_worksheet(nodes: list, output_wb: xl.Workbook):
    """
    Build a workbook object with data from summary nodes
    :param nodes: list of SummaryValue tuples
    :return: workbook object
    """
    summary_sheet = output_wb.create_sheet("summary")
    # write headers
    row = 1
    headers = ["Sheet Name", "Column Name", "Number of Differences", "Total Rows", "Percent Different", "Column Index"]
    for i in range(1, len(headers) + 1):
        format_header(summary_sheet.cell(row=row, column=i), headers[i - 1])

    # Write nodes
    for n in nodes:
        if isinstance(n, SummaryNode):
            row += 1
            node_values = [n.sheet_name, n.column_with_differences, n.number_of_differences, n.number_of_rows,
                           n.match_percent, n.column_index]
            for i in range(1, len(node_values) + 1):
                summary_sheet.cell(row=row, column=i).value = node_values[i - 1]

    # autosize columns
    for c in range(1, summary_sheet.max_column + 1):
        summary_sheet.column_dimensions[get_column_letter(c)].width = 30
    return output_wb


def format_header(cell, header_title):
    """
    Format and fill header cell
    :param cell: header cell
    :param header_title: header title
    :return: None
    """
    header_color = "A5FF00"
    header_pattern = PatternFill(start_color=header_color, fill_type="solid")
    cell.value = header_title
    cell.fill = header_pattern
