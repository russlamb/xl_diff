"""
This module is used to compare data files.  This project compares two excel or CSV files
and saves the results in an excel file.

Command line parsing has been moved to a separate file to keep code clean.

To run the module, you need to pass in three filenames: the two files being compared, denoted "left" and "right",
and the output file.  In addition to the filenames, there are several optional arguments to pass in that will alter
the comparison behavior.  By default the comparison will be a cell-by-cell comparison.  Other options include
Sorting based on a single column (e.g. a primary key) or group of columns (e.g. a composite key)

"""
import logging
import os
from collections import namedtuple

import openpyxl as xl
from dateutil.parser import parse
from openpyxl.styles import PatternFill

from .summary import create_summary_worksheet, get_workbook_nodes
from .validators import is_file_extension_valid, is_number, is_date

ValueNode = namedtuple('ValueNode', ['left_row', 'right_row', 'value'])  # object to store left row #, right #, value
SHEETS_PER_COMPARISON = 3


def make_sorted_sheet(workbook, sheet, sorted_values, new_sheet_name, left_or_right, has_header=True):
    """
    Create and sort excel worksheet according to the sorted_values.
    :param workbook: excel workbook object
    :param sheet: sheet object
    :param sorted_values: list of values by which to
    :param new_sheet_name: name of sheet
    :param left_or_right: indicates if sheet is left or right
    :param has_header: skip in sort
    :return: sorted excel sheet object
    """
    new_sheet = workbook.create_sheet(title=new_sheet_name)
    max_col = sheet.max_column

    logging.info("make sorted sheet: '{}', '{}'".format(new_sheet_name, left_or_right))
    if has_header:
        header_values = []
        for c in range(1, max_col + 1):
            header_values.append(sheet.cell(row=1, column=c).value)
        new_sheet.append(header_values)

    for v in sorted_values:
        if left_or_right == "left":
            row_id = v.left_row
        else:
            row_id = v.right_row

        # if a row id exists for this value & sheet, copy row data to new row in new sheet
        row_values = []
        if row_id is not None:  # if no row id, row_values will be empty
            for c in range(1, max_col + 1):
                row_values.append(sheet.cell(row=row_id, column=c).value)
        else:
            row_values = [None]
        new_sheet.append(row_values)  # append to new sheet
    return new_sheet


def copy_sheet_to_workbook(sheet, wb: xl.Workbook):
    """
    copy data from worksheet to a new worksheet in target workbook
    :param sheet: source worksheet
    :param wb: destination openpyxl workbook object
    :return: newly created openpyxl worksheet object in destination workbook
    """
    max_col = sheet.max_column
    max_row = sheet.max_row

    new_sheet = wb.create_sheet(sheet.title)
    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            new_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
    return new_sheet


def compare_files(left_path, right_path, output_path, threshold=0.001, open_on_finish=False, sort_column=None,
                  compare_type="default", has_header=True, sheet_matching="name", add_summary=True):
    """
    Compare two files and save comparison file.  Numerical differences that are below the threshold count as identical.
    Compare type indicates if rows are sorted or left alone.  Sheets are compared either by matching name or position.
    :param add_summary: if true, add a summary sheet to the output file with count of differences by column
    :param left_path: first file to compare.  Results show on left.
    :param right_path: second file to compare.  Results show on right
    :param output_path: output file.
    :param threshold: maximum acceptable differrnces of numerical values
    :param open_on_finish: if true, the output file will be opened when it is complete
    :param sort_column: numerical index of column, or list of such indices, used to sort rows
    :param compare_type: sorted or default (unsorted)
    :param has_header: if true, first row is excluded from sort
    :param sheet_matching: if "name", sheets with the same name will be compared.  Otherwise, order is used.
    :return: None
    """
    logging.info(
        "Comparing '{}' vs '{}' with threshold = '{}', sort column = '{}', compare type='{}'".format(
            left_path, right_path, threshold, sort_column, compare_type))

    # check file extension if valid.  If not, convert CSV to
    logging.info("validating file types: '{}', '{}'".format(left_path, right_path))
    left_path = is_file_extension_valid(left_path)
    right_path = is_file_extension_valid(right_path)

    # load workbook into excel library
    logging.info("loading excel files: '{}', '{}'".format(left_path, right_path))
    left_wb = xl.load_workbook(filename=left_path)
    right_wb = xl.load_workbook(filename=right_path)
    output_wb = xl.Workbook()

    # get sheet names
    logging.info("get sheet names")
    left_sheets = left_wb.sheetnames
    right_sheets = right_wb.sheetnames
    output_sheets = output_wb.sheetnames

    if sheet_matching == "name":
        sheets_to_process = [(sheet, sheet) for sheet in left_sheets if sheet in right_sheets]
    else:
        sheets_to_process = list(map(lambda i, j: (i, j), left_sheets, right_sheets))

    logging.info("sheet match style: '{}', sheets to process: {}".format(sheet_matching, sheets_to_process))

    if len(sheets_to_process) > 0:  # remove default sheet
        for sheet in output_sheets:
            output_wb.remove(output_wb[sheet])
    else:
        raise ValueError("No sheets were found for processing.  Check sheet_matching parameter is set correctly " +
                         "(e.g. name or order)")

    for (i, j) in sheets_to_process:
        left_sheet = left_wb[i]
        right_sheet = right_wb[j]

        if compare_type == "sorted":
            logging.info("sorting sheets prior to comparison: ({},{})".format(i, j))

            sorted_values = sort_values(left_sheet, right_sheet, sort_column, has_header)
            logging.info("values sorted.  Sorting left sheet")
            left_sheet = make_sorted_sheet(output_wb, left_sheet, sorted_values, 'left_' + i, 'left', has_header)
            logging.info("left sorted.  Sorting right sheet")
            right_sheet = make_sorted_sheet(output_wb, right_sheet, sorted_values, 'right_' + j, 'right',
                                            has_header)
        else:
            copy_sheet_to_workbook(left_sheet, output_wb)
            copy_sheet_to_workbook(right_sheet, output_wb)

        output_sheet_name = i if sheet_matching == "name" or i == j else "{} v {}".format(i, j)
        output_sheet = output_wb.create_sheet(output_sheet_name)

        logging.info("comparing sheets: ({},{})".format(i, j))
        compare_sheet(left_sheet, right_sheet, output_sheet, threshold)

    if add_summary:
        logging.info(f"worksheets {output_wb.worksheets}")
        logging.info("add summary sheet")
        workbook_nodes = get_workbook_nodes(SHEETS_PER_COMPARISON, output_wb)  # get differences for workbook
        logging.info(f"number of nodes {len(workbook_nodes)}")
        output_wb = create_summary_worksheet(workbook_nodes, output_wb)

    logging.info("saving to file: '{}'".format(output_path))
    output_wb.save(output_path)

    logging.info("save complete")
    if open_on_finish:
        path_to_open = '"' + output_path + '"'
        logging.info("opening file".format(path_to_open))
        os.system(path_to_open)  # use OS command line to open file.  This works on Windows


def get_list_of_values(row_number, sheet, sort_column):
    """
    returns a list of values from sheet specified by row number and sort_column(s).  If sort_column is a list,
    the result will be a list.
    :param row_number: row number
    :param sheet: sheet object
    :param sort_column: number or list of numbers indicating columns used for sorting
    :return: list of values from row and column coordinates.
    """
    try:
        return [sheet.cell(row=row_number, column=col).value if is_number(
            sheet.cell(row=row_number, column=col).value) else str(sheet.cell(row=row_number, column=col).value) for col
                in sort_column]
    except Exception as e:
        logging.error(f"row number: {row_number}, sort_column: {sort_column}, error: {e}")
        raise e


def sort_values(left, right, sort_column, has_header=False):
    """
    Line up values from left and right sheets for sorting.  Functions as a full outer join of the two data set indexes.
    To do this, we use a modified merge-sort algorithm.  Values sorted are tuples in case multiple columns are used.
    :param left: first worksheet object
    :param right: second worksheet object
    :param sort_column: number or list of numbers indicating columns used for sorting
    :param has_header:if true, first row is excluded from sort
    :return:sorted list of tuples indicating which rows from each sheet matches the value.
    """
    if sort_column is None:  # e.g., if not none
        return

    starting_row = 1 if has_header is False else 2

    # get list of named tuples. left side populates x.  right populates y.  merge later.
    if isinstance(sort_column, list):
        x = [ValueNode(n, None, tuple(get_list_of_values(n, left, sort_column))) for n in
             range(starting_row, left.max_row + 1)]  # x = left
        y = [ValueNode(None, n, tuple(get_list_of_values(n, right, sort_column))) for n in
             range(starting_row, right.max_row + 1)]  # y = right

        # get dictionary of columns and whether or not to convert them.  flag based on whether or not any item in the tuple is not a number.
        convert_column = [(col, (len([i for i in x if not is_number(i.value[col])]) > 0 or
                                 len([i for i in y if not is_number(i.value[col])]) > 0)) for col in
                          range(0, len(sort_column))]
        # convert

        x = sorted([ValueNode(i.left_row, i.right_row,
                              tuple([str(i.value[col]) if convert else i.value[col] for (col, convert) in
                                     convert_column])) for
                    i in x], key=lambda tup: tup[2])
        y = sorted([ValueNode(i.left_row, i.right_row,
                              tuple([str(i.value[col]) if convert else i.value[col] for (col, convert) in
                                     convert_column])) for
                    i in y], key=lambda tup: tup[2])


    else:
        x = [ValueNode(n, None, left.cell(row=n, column=sort_column).value) for n in
             range(starting_row, left.max_row + 1)]  # x = left
        y = [ValueNode(None, n, right.cell(row=n, column=sort_column).value) for n in
             range(starting_row, right.max_row + 1)]  # y = right

        # if there are any values that are not numbers, convert all to string
        if (
                len([i for i in x if not is_number(i.value)]) > 0 or
                len([i for i in y if not is_number(i.value)]) > 0):
            x = sorted([ValueNode(i.left_row, i.right_row, str(i.value)) for i in x], key=lambda tup: tup[2])
            y = sorted([ValueNode(i.left_row, i.right_row, str(i.value)) for i in y], key=lambda tup: tup[2])
        else:  # otherwise, convert all to number
            x = sorted([ValueNode(i.left_row, i.right_row, float(i.value)) for i in x], key=lambda tup: tup[2])
            y = sorted([ValueNode(i.left_row, i.right_row, float(i.value)) for i in y], key=lambda tup: tup[2])

    logging.debug("starting_row for sort: {}".format(starting_row))

    i = j = 0

    z = []  # z is combined list
    while i < len(x) and j < len(y):

        f = x[i].value
        r = y[j].value

        # compare left and right values.  if both values match, combine into a single tuple.  otherwise, take only one
        if f == r:
            temp = ValueNode(x[i].left_row, y[j].right_row, f)
            z.append(temp)
            i += 1
            j += 1
        elif f < r:  # left side has number lower than right side
            z.append(x[i])  # add node from left side since right side is None
            i += 1
        elif r < f:  # right side has number lower than left side
            z.append(y[j])  # add node from right side since left is None
            j += 1
        else:
            logging.warning("encountered a case where no comparison can be made {} vs {}".format(f, r))

    while i < len(x):
        z.append(x[i])
        i += 1
    while j < len(y):
        z.append(y[j])
        j += 1

    return z


def compare_sheet(left_sheet, right_sheet, output_sheet, threshold):
    """
    Compare two excel sheet objects.  Return output sheet.
    :param left_sheet: first sheet to compare (left)
    :param right_sheet: second sheet to compare (right)
    :param output_sheet: resulting sheet object
    :param threshold: numerical differences below this amount are considered identical
    :return: output sheet object containing comparison
    """
    max_col = max(left_sheet.max_column, right_sheet.max_column)
    max_row = max(left_sheet.max_row, right_sheet.max_row)

    columns_per_value = 3  # each comparison takes up 3 rows
    left_offset = 0
    right_offset = 1
    diff_offset = 2
    starting_row = 1

    for col in range(1, max_col + 1):
        output_column = (col - 1) * columns_per_value + 1  # 1-based column count, offset by columns per value
        for row in range(starting_row, max_row + 1):
            left_cell = left_sheet.cell(row=row, column=col)
            right_cell = right_sheet.cell(row=row, column=col)
            diff_value = cell_difference(left_cell, right_cell)
            output_sheet.cell(row=row, column=output_column + left_offset).value = left_cell.value  # output left
            output_sheet.cell(row=row, column=output_column + right_offset).value = right_cell.value  # output right
            output_sheet.cell(row=row, column=output_column + diff_offset).value = diff_value  # output diff
            apply_style(output_sheet.cell(row=row, column=output_column + diff_offset), threshold)

    return output_sheet


def cell_difference(left_cell, right_cell):
    """
    Compare two cell objects and get output comparison text
    :param left_cell: first cell to compare (left)
    :param right_cell:  seconds cell to compare(right)
    :return: string or number depending on the value types in the cells
    """
    return value_difference(left_cell.value, right_cell.value)


def value_difference(left, right):
    """
    Determine the difference between two values and return.  if both values are numbers, return a number.  If one is
    a string, return either Same or Different.
    :param left: first value to compare (left)
    :param right: second value to compare (right)
    :return: right - left, if both are numbers, or string indicating Sameness
    """
    if is_number(left) and is_number(right):
        diff_value = float(right) - float(left)  # numbers are subtracted
    elif is_date(left) and is_date(right):
        diff_value = "Same" if parse(right) == parse(left) else "Different"  # date comparison
    else:
        diff_value = "Same" if right == left else "Different"  # non-number comparison
    return diff_value


def apply_style(cell, threshold):
    """
    Apply style to difference cell based on contents.
    :param cell: cell containing difference
    :param threshold: differences below threshold are considered identical
    :return: None
    """
    same_color = "93f277"
    different_color = "edb26f"
    pattern_same = PatternFill(start_color=same_color, fill_type="solid")
    pattern_diferent = PatternFill(start_color=different_color, fill_type="solid")
    if is_number(cell.value) and abs(cell.value) <= threshold:
        cell.fill = pattern_same  # under threshold
    elif cell.value == "Same":
        cell.fill = pattern_same  # match
    else:
        cell.fill = pattern_diferent  # under threshold or different